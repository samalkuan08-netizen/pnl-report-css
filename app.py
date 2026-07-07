# -*- coding: utf-8 -*-
import io, tempfile, os
import streamlit as st
import openpyxl
import css_agent as eng
import recalc_native

st.set_page_config(page_title='P&L агент — CSS', layout='centered')
st.title('📊 P&L агент — филиал CSS (морские проекты)')
st.caption('Загрузка 1С ОСВ → заполнение единого листа P&L по блокам проектов. '
           'Новые статьи расходов >300 000 ₸ добавляются отдельной строкой автоматически.')

MONTHS={'Январь':'C','Февраль':'D','Март':'E','Апрель':'F','Май':'G','Июнь':'H',
        'Июль':'I','Август':'J','Сентябрь':'K','Октябрь':'L','Ноябрь':'M','Декабрь':'N'}

osv_files=st.file_uploader('Выгрузки ОСВ из 1С (можно несколько — по одной на месяц)',
                           type=['xlsx'], accept_multiple_files=True)
col1,col2=st.columns(2)
with col1:
    month=st.selectbox('Месяц (для одной выгрузки)', list(MONTHS.keys()))
base_file=st.file_uploader('Текущий отчёт P&L (чтобы дозаполнить новый месяц, не потеряв прежние)',
                           type=['xlsx'])

st.divider()

def _clean_month(ws, col):
    """Очистить колонку месяца от старых значений (кроме формул) — защита от накопления."""
    idx=openpyxl.utils.column_index_from_string(col)
    for r in range(3, ws.max_row+1):
        c=ws.cell(row=r,column=idx)
        if not (isinstance(c.value,str) and str(c.value).startswith('=')):
            c.value=None

if st.button('▶️ Сформировать отчёт', type='primary', disabled=not osv_files):
    try:
        # базовый шаблон: загруженный текущий отчёт или встроенный
        if base_file is not None:
            tmpl_bytes=base_file.getvalue()
        else:
            with open(os.path.join(os.path.dirname(__file__),'template.xlsx'),'rb') as f:
                tmpl_bytes=f.read()

        work=tempfile.NamedTemporaryFile(delete=False,suffix='.xlsx').name
        with open(work,'wb') as f: f.write(tmpl_bytes)

        # определить месяц для каждого файла: если файлов несколько — по порядку с выбранного
        start_idx=list(MONTHS.keys()).index(month)
        logs_all=[]; flags_all=[]
        for i,uf in enumerate(osv_files):
            m=list(MONTHS.keys())[min(start_idx+i,11)]; col=MONTHS[m]
            osv_tmp=tempfile.NamedTemporaryFile(delete=False,suffix='.xlsx').name
            with open(osv_tmp,'wb') as f: f.write(uf.getvalue())
            leaves=eng.parse_1c(osv_tmp)
            # очистка колонки месяца в текущем рабочем файле
            wb=openpyxl.load_workbook(work); ws=wb[eng.SHEET]; _clean_month(ws,col); wb.save(work)
            out=tempfile.NamedTemporaryFile(delete=False,suffix='.xlsx').name
            flags,log=eng.fill(work,leaves,out,month_col=col)
            work=out; flags_all+=[(m,)+tuple(x) for x in flags]; logs_all+=[(m,)+tuple(x) for x in log]
            st.info(f'✅ {uf.name} → колонка «{m}»: обработано статей {len(log)}')

        # пересчёт: формулы + кэш-значения (число видно сразу, формула сохраняется)
        recalc_native.recalc_native(work)

        with open(work,'rb') as f: data=f.read()
        st.success('Готово!')
        st.download_button('⬇️ Скачать отчёт P&L', data=data,
                           file_name='CSS_P_L_отчёт.xlsx',
                           mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        # новые статьи / флаги
        new_articles=[f for f in flags_all if len(f)>=5 and 'НОВАЯ' in str(f[4])]
        if new_articles:
            st.warning('🆕 Добавлены новые строки статей (>300 000 ₸):')
            for f in new_articles:
                st.write(f'• {f[0]}: «{f[2]}» в блоке — {f[3]:,.0f} ₸')
        other_flags=[f for f in flags_all if f not in new_articles]
        if other_flags:
            with st.expander(f'⚠️ Требуют внимания ({len(other_flags)})'):
                for f in other_flags: st.write('•', f)
    except Exception as e:
        st.error(f'Ошибка: {e}')
        import traceback; st.code(traceback.format_exc())

with st.expander('ℹ️ Как работает разноска'):
    st.markdown('''
- **Доход** (счета 6ххх) → блок «ДОХОД ОТ РЕАЛИЗАЦИИ» по проекту; вознаграждения (61хх) → «Доходы от финансовой деятельности»; курсовая (62хх) → «Прочие доходы».
- **Прямые затраты** (7010) → блок проекта (Зайсан, LQB, Flat Top, Jan De Nul, USDT, IPS…). Налоги по ЗП = ОПВ+ОСМС+соц.отчисления+соц.налог. «Аренда судна» → строка конкретного судна.
- **Административные** (7210/7212) → блок «Административные расходы» (командировочные = проезд+суточные+найм жилья, и т.д.).
- **Прочие расходы** (74хх, курсовые) → строка «Прочие расходы».
- **Новая статья**, которой нет в блоке: **>300 000 ₸** — добавляется отдельной строкой в блок; **≤300 000 ₸** — в «Прочие» блока.
- Доход тянется из 1С; ручные источники дохода добавляются бухгалтером отдельно.
''')
